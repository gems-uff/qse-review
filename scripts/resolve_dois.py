"""Step 0 – Extract papers from spreadsheets and resolve DOIs.

Scans ``papers/`` for all ``.xlsx`` files (or uses a path given via
``--input``), extracts all paper entries (year, authors, title, source
URL / DOI), and resolves a DOI for every entry that does not already
have one via:

  1. Regex extraction from the cell hyperlink URL (doi.org, dl.acm.org/doi/…).
  2. CrossRef title-search API (fallback, with respectful rate limiting).

Writes the result to ``out/dois.json`` (path overridable via ``--output``).
The file is **gitignored** — it lives in ``out/`` which is local-only.

This JSON is the canonical input for the next pipeline step
(``fetch_abstracts.py``) and replaces the old ``papers/*.pdf`` workflow
when working from a spreadsheet.

Output format (``out/dois.json``):
  A JSON array of objects, one per paper:
  {
    "sheet":    "<conference/journal name>",
    "year":     <int>,
    "authors":  "<author string>",
    "title":    "<paper title>",
    "doi":      "<DOI string or null>",
    "url":      "<fallback URL or null>",
    "doi_source": "spreadsheet" | "url_regex" | "crossref" | null
  }

Rate limiting:
  - CrossRef is called only when no DOI can be extracted from the URL.
  - Default delay between CrossRef calls: 1 s (--crossref-delay to adjust).
  - CrossRef ``score`` threshold: 30.0 (--min-score to adjust).
    Results below the threshold are discarded to avoid false matches.
"""

import argparse
import json
import logging
import re
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

try:
    import openpyxl  # type: ignore[import]
except ImportError:
    openpyxl = None  # type: ignore[assignment]

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
DEFAULT_PAPERS_DIR = Path("papers")
DEFAULT_OUTPUT = Path("out/dois.json")
CROSSREF_TIMEOUT = 15  # seconds
DBLP_TIMEOUT = 10  # seconds
DEFAULT_CROSSREF_DELAY = 1.0  # seconds between CrossRef calls
DBLP_DELAY = 2.0  # seconds between DBLP calls
DBLP_MAX_RETRIES = 3  # retries on 429 / connection errors
DEFAULT_MIN_SCORE = 30.0

# DOI regex: captures 10.XXXX/... from URLs with doi.org or dl.acm.org prefix
_DOI_RE = re.compile(
    r"(?:https?://(?:dx\.)?doi\.org/|https?://dl\.acm\.org/doi/)"
    r"(10\.\d{4,9}/[^\s,;\"'<>\[\]{}()]+)",
    re.IGNORECASE,
)

# DBLP record URL pattern
_DBLP_KEY_RE = re.compile(r"dblp\.org/rec/(.+)", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Spreadsheet parsing
# ---------------------------------------------------------------------------

def _parse_title_cell(cell) -> tuple[str | None, str | None, str | None]:
    """Return (doi, url, title) from a spreadsheet cell.

    Three formats are handled:
    - ``=HYPERLINK("url", "title")`` formula (stored as literal string)
    - External hyperlink object (``cell.hyperlink.target``)
    - Plain text title with no link
    """
    val = cell.value
    hl = cell.hyperlink

    # --- Formula style: =HYPERLINK("url", "title") -----------------------
    if isinstance(val, str) and val.strip().upper().startswith("=HYPERLINK"):
        m = re.match(r'=HYPERLINK\("([^"]+)",\s*"([^"]*)"\)', val.strip(), re.IGNORECASE)
        if m:
            link_url, title = m.group(1), m.group(2)
            doi = _doi_from_url(link_url)
            return doi, (None if doi else link_url), title
        # malformed formula — treat whole value as title
        return None, None, val

    # --- External hyperlink object ----------------------------------------
    title = val if isinstance(val, str) else None
    if hl and hl.target:
        doi = _doi_from_url(hl.target)
        return doi, (None if doi else hl.target), title

    return None, None, title


def _doi_from_url(url: str) -> str | None:
    """Extract and normalise a DOI from a URL, or return None."""
    m = _DOI_RE.search(url)
    if m:
        return m.group(1).rstrip(".")
    return None


def _parse_spreadsheet(path: Path) -> list[dict]:
    """Return a list of raw paper dicts from all sheets."""
    if openpyxl is None:
        raise ImportError(
            "openpyxl is required to read the spreadsheet.\n"
            "Install it with: pip install openpyxl"
        )

    wb = openpyxl.load_workbook(path)
    papers: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            # Expect columns: YEAR | AUTHORS | TITLE  (possibly offset by 1 header row)
            if len(row) < 3:
                continue
            year_cell, authors_cell, title_cell = row[0], row[1], row[2]

            year = year_cell.value
            if not isinstance(year, (int, float)):
                continue  # header or empty row
            year = int(year)

            authors = authors_cell.value
            if not isinstance(authors, str) or not authors.strip():
                continue

            doi, url, title = _parse_title_cell(title_cell)

            if not title or title.strip().upper() == "TITLE":
                continue

            papers.append(
                {
                    "sheet": sheet_name,
                    "year": year,
                    "authors": authors.strip(),
                    "title": title.strip(),
                    "doi": doi,
                    "url": url,
                    "doi_source": "spreadsheet" if doi else None,
                }
            )

    logger.info("Parsed %d papers from %d sheets.", len(papers), len(wb.sheetnames))
    return papers


# ---------------------------------------------------------------------------
# DBLP BibTeX fallback
# ---------------------------------------------------------------------------

def _dblp_resolve(dblp_url: str) -> str | None:
    """Fetch the BibTeX record for a dblp.org/rec/… URL and extract the DOI."""
    m = _DBLP_KEY_RE.search(dblp_url)
    if not m:
        return None
    key = m.group(1).rstrip("/")
    bib_url = f"https://dblp.org/rec/{key}.bib"
    req = urllib.request.Request(
        bib_url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (compatible; qse-review/1.0; "
                "+https://github.com/qse-review)"
            )
        },
    )
    bib_text: str | None = None
    for attempt in range(1, DBLP_MAX_RETRIES + 1):
        try:
            with urllib.request.urlopen(req, timeout=DBLP_TIMEOUT) as resp:
                bib_text = resp.read().decode("utf-8", errors="replace")
            break
        except urllib.error.HTTPError as exc:
            if exc.code == 429:
                wait = 2 ** attempt  # 2, 4, 8 seconds
                logger.warning(
                    "DBLP 429 for %s — backoff %ds (attempt %d/%d)",
                    key, wait, attempt, DBLP_MAX_RETRIES,
                )
                time.sleep(wait)
            else:
                logger.warning("DBLP BibTeX fetch failed for %s: %s", key, exc)
                return None
        except OSError as exc:
            wait = 2 ** attempt
            logger.warning(
                "DBLP connection error for %s: %s — backoff %ds (attempt %d/%d)",
                key, exc, wait, attempt, DBLP_MAX_RETRIES,
            )
            time.sleep(wait)
    if bib_text is None:
        logger.warning("DBLP: gave up after %d retries for %s", DBLP_MAX_RETRIES, key)
        return None

    doi_m = re.search(r'doi\s*=\s*[{"]([^}"]+)[}"]', bib_text, re.IGNORECASE)
    if doi_m:
        doi = doi_m.group(1).strip().rstrip(".")
        logger.debug("  DBLP resolved DOI %s for key %s", doi, key)
        return doi

    logger.debug("  DBLP BibTeX has no DOI field for key %s", key)
    return None


# ---------------------------------------------------------------------------
# CrossRef title-search fallback
# ---------------------------------------------------------------------------

def _crossref_resolve(title: str, min_score: float, mailto: str | None) -> str | None:
    """Query CrossRef by title and return a DOI if confidence is high enough."""
    query = urllib.parse.quote(title[:200])
    url = f"https://api.crossref.org/works?query.title={query}&rows=1&select=DOI,title,score"

    mailto_str = mailto or "qse-review@local"
    req = urllib.request.Request(
        url,
        headers={"User-Agent": f"qse-review/1.0 (mailto:{mailto_str})"},
    )
    try:
        with urllib.request.urlopen(req, timeout=CROSSREF_TIMEOUT) as resp:
            data = json.loads(resp.read())
    except (urllib.error.URLError, OSError, json.JSONDecodeError) as exc:
        logger.warning("CrossRef request failed for %r: %s", title[:60], exc)
        return None

    items = data.get("message", {}).get("items", [])
    if not items:
        return None

    item = items[0]
    score = item.get("score", 0.0)
    if score < min_score:
        logger.debug("  CrossRef score %.1f below threshold for %r", score, title[:60])
        return None

    doi = item.get("DOI")
    if doi:
        logger.debug("  CrossRef resolved DOI %s (score=%.1f)", doi, score)
    return doi


# ---------------------------------------------------------------------------
# Main resolution loop
# ---------------------------------------------------------------------------

def resolve_dois(
    papers: list[dict],
    crossref_delay: float,
    min_score: float,
    mailto: str | None,
    skip_crossref: bool,
) -> list[dict]:
    """Resolve missing DOIs via DBLP then CrossRef; return the updated list."""
    needs_resolution = [p for p in papers if p["doi"] is None]
    has_doi = len(papers) - len(needs_resolution)

    logger.info(
        "%d papers already have a DOI; %d need further resolution.",
        has_doi,
        len(needs_resolution),
    )

    if not needs_resolution:
        return papers

    # --- Pass 1: DBLP BibTeX (for entries linking to dblp.org/rec/…) ------
    dblp_candidates = [
        p for p in needs_resolution
        if p.get("url") and "dblp.org/rec/" in p["url"]
    ]
    if dblp_candidates:
        logger.info("Pass 1 – DBLP BibTeX resolution (%d entries)…", len(dblp_candidates))
        for i, paper in enumerate(dblp_candidates, 1):
            logger.info("  [%d/%d] %s", i, len(dblp_candidates), paper["title"][:70])
            doi = _dblp_resolve(paper["url"])
            if doi:
                paper["doi"] = doi
                paper["doi_source"] = "dblp"
            else:
                logger.warning("    → DBLP: no DOI found; URL kept: %s", paper["url"])
            if i < len(dblp_candidates):
                time.sleep(DBLP_DELAY)
        dblp_resolved = sum(1 for p in dblp_candidates if p["doi"])
        logger.info("DBLP done — resolved: %d / %d", dblp_resolved, len(dblp_candidates))

    # --- Pass 2: CrossRef by title (remaining entries without DOI) ---------
    if skip_crossref:
        return papers

    still_missing = [p for p in papers if p["doi"] is None]
    if not still_missing:
        return papers

    logger.info(
        "Pass 2 – CrossRef title resolution (%d entries, %.1fs delay)…",
        len(still_missing),
        crossref_delay,
    )
    for i, paper in enumerate(still_missing, 1):
        logger.info("  [%d/%d] %s", i, len(still_missing), paper["title"][:70])
        doi = _crossref_resolve(paper["title"], min_score, mailto)
        if doi:
            paper["doi"] = doi
            paper["doi_source"] = "crossref"
        else:
            logger.warning("    → CrossRef: no DOI resolved; URL kept: %s", paper.get("url"))
        if i < len(still_missing):
            time.sleep(crossref_delay)

    crossref_resolved = sum(1 for p in still_missing if p["doi"])
    logger.info(
        "CrossRef done — resolved: %d  still missing: %d",
        crossref_resolved,
        len(still_missing) - crossref_resolved,
    )
    return papers


# ---------------------------------------------------------------------------
# Unresolved report
# ---------------------------------------------------------------------------

def _write_unresolved_report(papers: list[dict], output_path: Path) -> None:
    """Write papers without a DOI to a separate JSON for manual follow-up."""
    unresolved = [p for p in papers if not p["doi"]]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as fh:
        json.dump(unresolved, fh, ensure_ascii=False, indent=2)
    if unresolved:
        logger.warning(
            "%d paper(s) without a DOI written to %s — manual follow-up required.",
            len(unresolved),
            output_path,
        )
    else:
        logger.info("All papers resolved — %s is empty.", output_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Extract papers from all spreadsheets in papers/ and resolve DOIs."
    )
    p.add_argument(
        "--input",
        type=Path,
        default=None,
        metavar="XLSX_OR_DIR",
        help=(
            "Path to a specific .xlsx file, or a directory to scan for all .xlsx files "
            f"(default: {DEFAULT_PAPERS_DIR})."
        ),
    )
    p.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        metavar="JSON",
        help=f"Path for the output JSON file (default: {DEFAULT_OUTPUT}; gitignored).",
    )
    p.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite the output file if it already exists.",
    )
    p.add_argument(
        "--skip-crossref",
        action="store_true",
        help="Do not call CrossRef; only use DOIs already present in the spreadsheet.",
    )
    p.add_argument(
        "--crossref-delay",
        type=float,
        default=DEFAULT_CROSSREF_DELAY,
        metavar="SECONDS",
        help=f"Seconds to wait between CrossRef API calls (default: {DEFAULT_CROSSREF_DELAY}).",
    )
    p.add_argument(
        "--min-score",
        type=float,
        default=DEFAULT_MIN_SCORE,
        metavar="SCORE",
        help=(
            f"Minimum CrossRef relevance score to accept a DOI match "
            f"(default: {DEFAULT_MIN_SCORE})."
        ),
    )
    p.add_argument(
        "--mailto",
        type=str,
        default=None,
        metavar="EMAIL",
        help=(
            "E-mail address for the CrossRef Polite Pool "
            "(higher rate limits). Optional but recommended."
        ),
    )
    return p


def _paper_key(paper: dict) -> tuple:
    """Stable identity key for deduplication: (sheet, normalised title)."""
    return (paper["sheet"], paper["title"].strip().lower())


def _load_cache(path: Path) -> dict:
    """Load existing output JSON as a dict keyed by _paper_key, or empty dict."""
    if not path.exists():
        return {}
    try:
        with open(path, encoding="utf-8") as fh:
            existing: list[dict] = json.load(fh)
        cache = {_paper_key(p): p for p in existing if isinstance(p, dict)}
        logger.info("Loaded %d cached entries from %s.", len(cache), path)
        return cache
    except (json.JSONDecodeError, OSError) as exc:
        logger.warning("Could not read cache %s: %s — starting fresh.", path, exc)
        return {}


def _find_xlsx_files(input_arg: Path | None) -> list[Path]:
    """Return the list of .xlsx files to process."""
    if input_arg is None:
        target = DEFAULT_PAPERS_DIR
    else:
        target = input_arg

    if target.is_file():
        return [target]

    if target.is_dir():
        xlsx_files = sorted(target.glob("*.xlsx"))
        if not xlsx_files:
            logger.error("No .xlsx files found in %s", target)
            raise SystemExit(1)
        logger.info("Found %d spreadsheet(s): %s", len(xlsx_files), [f.name for f in xlsx_files])
        return xlsx_files

    logger.error("Input path not found: %s", target)
    raise SystemExit(1)


def main() -> None:
    args = _build_parser().parse_args()

    xlsx_files = _find_xlsx_files(args.input)

    # Load the existing output as a cache (empty if --overwrite or file absent)
    cache = {} if args.overwrite else _load_cache(args.output)

    papers: list[dict] = []
    for xlsx_path in xlsx_files:
        logger.info("Parsing %s…", xlsx_path.name)
        papers.extend(_parse_spreadsheet(xlsx_path))

    # Apply cache: reuse already-resolved entries; mark new/unresolved ones
    new_count = cached_count = 0
    for paper in papers:
        key = _paper_key(paper)
        if key in cache and cache[key].get("doi"):
            # Reuse the cached resolved entry wholesale
            paper.update(cache[key])
            cached_count += 1
        else:
            new_count += 1

    logger.info(
        "Incremental mode — cached (reused): %d  new/unresolved: %d",
        cached_count,
        new_count,
    )

    papers = resolve_dois(
        papers,
        crossref_delay=args.crossref_delay,
        min_score=args.min_score,
        mailto=args.mailto,
        skip_crossref=args.skip_crossref,
    )

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as fh:
        json.dump(papers, fh, ensure_ascii=False, indent=2)

    total = len(papers)
    with_doi = sum(1 for p in papers if p["doi"])
    logger.info(
        "Wrote %d papers to %s  (with DOI: %d / %d)",
        total,
        args.output,
        with_doi,
        total,
    )

    unresolved_path = args.output.parent / "unresolved_papers.json"
    _write_unresolved_report(papers, unresolved_path)


if __name__ == "__main__":
    main()
